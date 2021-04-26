print('Hello, world!')
print(1 + 2)

import openpyxl
  wb = openpyxl.load_workbook("Book1.xlsx")

from pathlib import Path
import openpyxl
p = Path("./Book1.xlsx")
if p.exists():
  wb = openpyxl.load_workbook(p)

import openpyxl
  wb = openpyxl.load_workbook("Book1.xlsm", keep_vba=True)

import openpyxl
wb = openpyxl.Workbook()